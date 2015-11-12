try:
    from collections import OrderedDict
except ImportError:
    from ordereddict import OrderedDict
import os
import sys
import time
import click
import polib
import openpyxl
from . import ColumnHeaders
try:
    unicode
except NameError:
    unicode = str


def save(output_file, catalog):
    """Save catalog to a PO file.

    This is mostly a stripped down copy of POFile.save so we can save the
    catalog to a file safely created by click.
    """
    output_file.write(unicode(catalog))


def po_timestamp(filename):
    local = time.localtime(os.stat(filename).st_mtime)
    offset = -(time.altzone if local.tm_isdst else time.timezone)
    return '%s%s%s' % (
        time.strftime('%Y-%m-%d %H:%M', local),
        '-' if offset < 0 else '+',
        time.strftime('%H%M', time.gmtime(abs(offset))))


@click.command()
@click.argument('locale', required=True)
@click.argument('input_file',
        type=click.Path(exists=True, readable=True),
        required=True)
@click.argument('output_file', type=click.File('w', encoding='utf-8'), required=True)
def main(locale, input_file, output_file):
    """
    Convert a XLS(X) file to a .PO file
    """
    book = openpyxl.load_workbook(input_file)
    catalog = polib.POFile()
    catalog.header = u'This file was generated from %s' % input_file
    catalog.metata_is_fuzzy = True
    catalog.metadata = OrderedDict()
    catalog.metadata['PO-Revision-Date'] = po_timestamp(input_file)
    catalog.metadata['Content-Type'] = 'text/plain; charset=UTF-8'
    catalog.metadata['Content-Transfer-Encoding'] = '8bit'
    catalog.metadata['Generated-By'] = 'xls-to-po 1.0'

    for sheet in book.worksheets:
        if sheet.max_row < 2:
            continue
        click.echo('Processing sheet %s' % sheet.title)
        row_iterator = sheet.iter_rows()
        headers = [c.value for c in next(row_iterator)]
        headers = dict((b, a) for (a, b) in enumerate(headers))
        msgctxt_column = headers.get(ColumnHeaders.msgctxt)
        msgid_column = headers.get(ColumnHeaders.msgid)
        tcomment_column = headers.get(ColumnHeaders.tcomment)
        msgstr_column = headers.get(locale)
        if msgid_column is None:
            click.echo(u'Could not find a "%s" column' % ColumnHeaders.msgid,
                    err=True)
            continue
        if msgstr_column is None:
            click.echo(u'Could not find a "%s" column' % locale, err=True)
            continue

        with click.progressbar(row_iterator, length=sheet.max_row - 1,
                label='Extracting messages') as rows:
            for row in rows:
                row = [c.value for c in row]
                if not row[msgid_column]:
                    continue
                try:
                    entry = polib.POEntry(
                            msgid=row[msgid_column],
                            msgstr=row[msgstr_column] or '')
                    if msgctxt_column is not None and row[msgctxt_column]:
                        entry.msgctxt = row[msgctxt_column]
                    if tcomment_column:
                        entry.tcomment = row[tcomment_column]
                    catalog.append(entry)
                except IndexError:
                    click.echo('Row %s is too short' % row, err=True)

    if not catalog:
        click.echo('No messages found, aborting', err=True)
        sys.exit(1)

    save(output_file, catalog)


if __name__ == '__main__':
    main()
